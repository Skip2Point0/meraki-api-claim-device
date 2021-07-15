import requests
import json
from openpyxl import load_workbook

# ################################################# PARAMETERS #########################################################
meraki_api = 'APIKEY'
organization_id = 'Organization Name'
network_ids = ['Network Name1', 'Network Name2']
spread = 'Spreadsheet.xlsx'
tabs = ['Sheet Name1', 'Sheet Name2']
address = ['Physical Address of First Network', 'Physical Address of Second Network']
ap_abr = ['NN1-', 'NN2-']
ap_name_column = 'A'
notes_column = 'B'
tags_column = 'C'
serials_column = 'U'
mac_column = 'V'
# ################################################# PARAMETERS #########################################################

net_dictionary = {}
headers = {
    'X-Cisco-Meraki-API-Key': meraki_api,
    'Content-Type': 'application/json'
}

print(spread)
wb = load_workbook(spread)


def pull_organization_id(head):
    url = "https://api.meraki.com/api/v0/organizations"
    payload = {}

    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    response = json.loads(response)
    for dicti in response:
        name = dicti["name"]
        if name == organization_id:
            org_id = dicti["id"]
            print("#################################################")
            print(name + "\n" + "Organization ID: " + org_id)
            print("#################################################")
            return org_id
        else:
            continue


def pull_organization_networks(head):
    global net_dictionary
    global organization_id

    organization_id = pull_organization_id(head)
    url = "https://api.meraki.com/api/v0/organizations/" + organization_id + "/networks"
    payload = {}
    response = requests.request("GET", url, headers=head, data=payload)
    response = response.content
    json_response = json.loads(response)
    for networks in json_response:
        name = networks['name']
        n_id = networks['id']
        net_dictionary[name] = n_id
    print(net_dictionary)
    return net_dictionary


def pull_destination_networks():
    global network_ids

    dest_network_ids = []
    for n in network_ids:
        for i in net_dictionary:
            if n == i:
                print("Destination Network: " + n)
                dest_network_ids.append(net_dictionary[n])
                break
            else:
                continue
    print(dest_network_ids)
    return dest_network_ids


pull_organization_networks(headers)
networks_dest = pull_destination_networks()


def meraki_claim_serial(workbook, networks, tbs, abr, serial, name, head):
    ap_index = []
    print("########################################")
    print("CLAIMING SERIAL NUMBER...")
    incr = 0
    for tab in tbs:
        sheet = workbook[tab]
        row_index = []
        print("########################################")
        print("NETWORK:  " + tab)
        print("########################################")
        print(networks)
        incr2 = 1
        for row in range(sheet.max_row):
            a = sheet[name + str(incr2)]
            unsorted_ap = a.value
            incr2 = incr2 + 1

            if abr[incr] in str(unsorted_ap):
                n_ap_index = row + 1
                ap_index.append(unsorted_ap)
                row_index.append(row)
                serials = sheet[serial + str(n_ap_index)].value
                url = ("https://api.meraki.com/api/v0/networks/" + networks[incr] + "/devices/claim")
                print(url)
                payload = {
                    "serial": serials,
                }
                payload = json.dumps(payload)
                print(payload)
                response = requests.request("POST", url, headers=head, data=payload)
                print(response.text.encode('utf8'))

        incr = incr + 1


meraki_claim_serial(wb, networks_dest, tabs, ap_abr, serials_column, ap_name_column, headers)